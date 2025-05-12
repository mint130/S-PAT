import asyncio
import base64
import io
import json
import logging
import os
import re
import time
from typing import Any, AsyncGenerator, Dict, List, Union
from fastapi import APIRouter, BackgroundTasks, Depends, File, HTTPException, UploadFile, status
from fastapi.responses import FileResponse, JSONResponse, StreamingResponse
from langchain_google_genai import ChatGoogleGenerativeAI
import pandas as pd
from sqlalchemy.orm import Session
from pydantic import BaseModel
from datetime import datetime
from dotenv import load_dotenv
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from langchain_openai import ChatOpenAI
from langchain.memory import ConversationBufferMemory
from langchain_community.chat_message_histories import ChatMessageHistory
from langchain_core.prompts import ChatPromptTemplate
from langchain_core.output_parsers import JsonOutputParser
from langchain.output_parsers import PydanticOutputParser
from langchain_core.output_parsers import StrOutputParser
from langchain_core.runnables import RunnablePassthrough
from langchain.schema.runnable import RunnableMap

from langchain.text_splitter import RecursiveCharacterTextSplitter
from langchain_openai import OpenAIEmbeddings
from langchain_community.vectorstores import FAISS
from langchain.chains import RetrievalQA
from langchain_xai import ChatXAI
from langchain_anthropic import ChatAnthropic

from app.api.endpoints.user import classify_patent, create_faiss_database, process_standards_for_vectordb
from app.core.redis import get_redis
from app.schemas.classification import ClassificationResponse, ClassificationSchema, Patent
from app.schemas.conversation import ConversationRequest, ConversationResponse, SessionHistoryResponse
from app.db.database import get_db
from app.models.conversation import Conversation
from app.crud.crud_conversation import create_conversation_record, get_conversation_history_by_session
from app.schemas.test import Progress, Message


load_dotenv()

# 로그
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# 결과 저장 디렉토리
RESULT_DIR = "./classified_excels"
os.makedirs(RESULT_DIR, exist_ok=True)

test_router = APIRouter(prefix="/test")
embeddings_openai = OpenAIEmbeddings(model="text-embedding-3-large", dimensions=3072)

gpt = ChatOpenAI(api_key=os.getenv("OPENAI_API_KEY"), model_name="gpt-4o", temperature=0)
claude = ChatAnthropic(
    model="claude-3-7-sonnet-20250219",
    temperature=0,
    api_key=os.getenv("CLAUDE_API_KEY")
)
gemini = ChatGoogleGenerativeAI(
    model="gemini-2.0-flash",
    temperature=0,
    google_api_key=os.getenv("GEMINI_API_KEY")
)
grok = ChatXAI(
    model="grok-3-beta",
    temperature=0,
    xai_api_key=os.getenv("GROK3_API_KEY")
)

user_vector_stores: Dict[str, FAISS] = {}

# 세션별 작업 진행 큐를 저장할 딕셔너리
session_progress_queues: Dict[str, asyncio.Queue] = {}

# SSE 이벤트 생성 함수
def format_sse(data: Union[str, dict], event=None) -> str:
    if isinstance(data, dict):
        data = json.dumps(data, ensure_ascii=False)
    msg = f"data: {data}\n\n"
    if event is not None:
        msg = f"event: {event}\n{msg}"
    return msg

# 특허 분류 작업을 백그라운드에서 실행하는 함수
async def process_patent_classification(
    session_id: str,
    llm,
    retriever,
    df: pd.DataFrame,
    redis,
    progress_queue: asyncio.Queue
):
    logger.info(f"호출 완료")
    try:
        logger.info(f"시작")
        patents = []
        total_patents = len(df)
        
        # 로그 추가
        start_time = datetime.now()
        logger.info(f"[{session_id}] 분류 시작 시간: {start_time}")
        
        # 각 행에 대해 RAG 처리 및 분류 추가
        for index, row in df.iterrows():
            # 출원번호와 특허 정보 확인
            application_number = row.get('출원번호', f"KR10-XXXX-{index:07d}")  # 출원번호가 없을 경우 임의 생성
            title = row.get('특허명', row.get('발명의 명칭', ''))
            abstract = row.get('요약', '')
            
            # 특허 정보가 없으면 다음 행으로
            if not title and not abstract:
                continue
                
            # 특허 정보 구성
            patent_info = f"특허명: {title} 요약: {abstract}"
            
            # RAG를 통한 분류
            classifications = await classify_patent(llm, retriever, patent_info)
            
            # 원본 데이터프레임에 분류 결과 추가
            df.loc[index, '대분류코드'] = classifications["majorCode"]
            df.loc[index, '중분류코드'] = classifications["middleCode"]
            df.loc[index, '소분류코드'] = classifications["smallCode"]
            df.loc[index, '대분류명칭'] = classifications["majorTitle"]
            df.loc[index, '중분류명칭'] = classifications["middleTitle"]
            df.loc[index, '소분류명칭'] = classifications["smallTitle"]
            
            # 특허 정보 구성 (요청된 JSON 형식에 맞게)
            patent_data = Patent(
                applicationNumber=application_number,
                title=title,
                abstract=abstract,
                majorCode=classifications["majorCode"],
                middleCode=classifications["middleCode"],
                smallCode=classifications["smallCode"],
                majorTitle=classifications["majorTitle"],
                middleTitle=classifications["middleTitle"],
                smallTitle=classifications["smallTitle"]
            )
            
            # 1. 로그 출력
            logger.info(f"[{session_id}] 분류된 특허: {patent_data.model_dump()}")

            # 2. redis에 저장
            redis.rpush(session_id, patent_data.model_dump_json())

            # 3. 리스트 추가
            patents.append(patent_data)
            
            # 4. 진행률 업데이트
            progress_data = Progress(
                current=index + 1,
                total=total_patents,
                percentage=round(((index + 1) / total_patents) * 100, 2),
            )
            await progress_queue.put(progress_data.model_dump())  

            # time.sleep(0.5)
            await asyncio.sleep(0.5)
            
        # 세션 데이터 만료 시간 설정 (24시간)
        redis.expire(session_id, 86400)  # 86400초 = 24시간
        
        # 종료 시간 기록
        end_time = datetime.now()
        elapsed = end_time - start_time
        logger.info(f"[{session_id}] 분류 종료 시간: {end_time}")
        logger.info(f"[{session_id}] 분류 소요 시간: {elapsed} (총 {elapsed.total_seconds():.2f}초)")

        # 임시 파일명 생성
        filename = f"{session_id}_classified.xlsx"
        save_path = f"./classified_excels/{filename}"

        # 저장 디렉토리 확인
        os.makedirs(os.path.dirname(save_path), exist_ok=True)

        # 먼저 데이터프레임을 엑셀로 저장
        buffer = io.BytesIO()
        with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name="Sheet1")
        buffer.seek(0)

        # 헤더 노란색 처리
        wb = load_workbook(buffer)
        ws = wb.active
        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        for cell in ws[1]:
            cell.fill = yellow_fill

        # 저장
        wb.save(save_path)
        
        # 작업 완료 메시지
        completion_data = Message(
            status="completed",
            message="분류 작업이 완료되었습니다.",
        )

        await progress_queue.put(completion_data.model_dump())  
        
    except Exception as e:
        # 오류가 발생한 경우 메시지 전송
        
        error_message = Message(
            status="error",
            message=f"오류가 발생했습니다: {str(e)}"
        )
       
        await progress_queue.put(error_message.model_dump())

# SSE 스트리밍 생성기
async def progress_event_generator(progress_queue: asyncio.Queue) -> AsyncGenerator[str, None]:
    try:
        while True:
            # 큐에서 메시지를 가져옴
            message = await progress_queue.get()
            
            # SSE 형식으로 메시지 전송
            event_data = format_sse(message)
            yield event_data
            
            # 완료 또는 오류 메시지인 경우 종료
            if isinstance(message, str):
                data = json.loads(message)
            else:
                data = message  # 이미 dict인 경우 그대로 사용
            
            # 오류인 경우 멈춤 
            if data.get("status") == "error":
                break

            # 완료된 경우 연결 종료
            if data.get("status") == "completed":
                yield format_sse("done", event="done")
                break
    except asyncio.CancelledError:
        # 클라이언트 연결 종료 시
        logger.info("클라이언트 연결이 종료되었습니다.")

    
@test_router.post("/{session_id}/standard/save", status_code=status.HTTP_201_CREATED, summary="분류 체계 벡터 db에 저장", description="선택된 분류 체계를 Redis 및 FAISS에 저장")
async def save_standard(session_id: str, standard: ConversationResponse, db:Session = Depends(get_db)):
    
    standards = standard.standards
    
    if not standards:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="분류 체계 데이터가 없습니다."
        )
    
    # 벡터 DB에 저장할 문서 준비
    documents = process_standards_for_vectordb(standards)
    
    # FAISS 벡터 DB 생성 및 메모리에 저장
    vectordb = create_faiss_database(documents, session_id)
    # print("벡터 db", vectordb.docstore._dict)
    
    # 전역 딕셔너리에 저장
    user_vector_stores[session_id] = vectordb


    return True
    

# 1. 파일 업로드 엔드포인트 (작업 시작만 담당), 202 반환
@test_router.post("/{session_id}/upload", status_code=status.HTTP_202_ACCEPTED, response_model = Message, summary="특허 데이터 파일 업로드", description="분류 하고 싶은 특허 데이터 파일을 업로드합니다.")
async def upload_and_start_classification(
    background_tasks: BackgroundTasks,
    session_id: str, 
    LLM: str, 
    file: UploadFile = File(...),
    redis = Depends(get_redis)
):
    # 업로드된 파일이 엑셀 파일인지 확인
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="업로드된 파일은 Excel (.xlsx 또는 .xls) 형식이어야 합니다")
    
    # session_id에 대한 벡터 스토어가 존재하는지 확인
    if session_id not in user_vector_stores:
        raise HTTPException(
            status_code=404, 
            detail="해당 세션에 대한 분류 체계가 존재하지 않습니다. 먼저 분류 체계를 저장해주세요."
        )

    if len(user_vector_stores[session_id].docstore._dict) == 0:
        raise HTTPException(status_code=400, detail="벡터 DB에 분류 체계 데이터가 없습니다. 먼저 분류 체계를 저장해주세요.")

    try:
        # 최적의 LLM 셋팅
        if LLM=="GPT":
            llm = gpt
        elif LLM == "CLAUDE":
            llm = claude
        elif LLM == "GEMINI":
            llm = gemini
        elif LLM == "GROK":
            llm = grok

        # 파일 내용을 메모리에서 읽기
        contents = await file.read()
        
        # 바이트 스트림을 IO 객체로 변환
        file_object = io.BytesIO(contents)
        
        # 엑셀 파일을 데이터프레임으로 읽기
        df = pd.read_excel(file_object)
        
        # retriever 생성
        retriever = user_vector_stores[session_id].as_retriever(search_kwargs={"k": 3})
        
        # 진행률 통신을 위한 큐 생성
        # 이미 존재하는 큐가 있다면 제거
        if session_id in session_progress_queues:
            del session_progress_queues[session_id]
        
        # 새 큐 생성
        progress_queue = asyncio.Queue()
        session_progress_queues[session_id] = progress_queue
        
        # 백그라운드에서 분류 작업 실행
        asyncio.create_task(
            process_patent_classification(
                session_id,
                llm,
                retriever,
                df,
                redis,
                progress_queue
            )
        )

        message = Message(
            status="processing",
            message="분류 작업이 시작되었습니다."
        )
        
        # 작업 시작 성공 응답 반환
        return message
        
    except Exception as e:
        # 예외 처리
        logger.error(f"분류 작업 시작 중 오류 발생: {str(e)}")
        raise HTTPException(status_code=500, detail=f"분류 작업 시작 중 오류가 발생했습니다: {str(e)}")


# 2. 진행 상황 스트리밍 엔드포인트 (SSE)
@test_router.get("/{session_id}/progress", summary="특허 분류 진행도 반환", description="특허 분류 진행도를 퍼센테이지로 반환합니다.")
async def stream_classification_progress(session_id: str):
    # 세션에 대한 진행 큐가 존재하는지 확인
    if session_id not in session_progress_queues:
        raise HTTPException(
            status_code=404,
            detail="해당 세션에 대한 진행 중인 분류 작업을 찾을 수 없습니다."
        )
    
    # 해당 세션의 진행 큐 가져오기
    progress_queue = session_progress_queues[session_id]
    
    # SSE 스트림 응답 반환
    return StreamingResponse(
        progress_event_generator(progress_queue),
        media_type="text/event-stream",
        headers={"Content-Type": "text/event-stream; charset=utf-8"}
    )

@test_router.get("/{session_id}/classification", summary="특허 분류 결과 json 반환", response_model=ClassificationResponse, description="세션 아이디로 저장된 분류결과 JSON을 반환합니다.")
async def get_classified_patents(session_id: str, redis = Depends(get_redis)):
    # redis에 저장된 세션이 없을 경우
    if not redis.exists(session_id):
        raise HTTPException(status_code=404, detail="해당 세션에 대한 분류 결과가 존재하지 않습니다.")

    items = redis.lrange(session_id, 0, -1)
    patents = [Patent.model_validate(json.loads(item)) for item in items]
    return ClassificationResponse(patents=patents)


@test_router.get("/{session_id}/classification/excel", summary="특허 분류 결과 엑셀 파일 반환", description="사용자 로컬 디스크에 저장된 분류 결과 엑셀 파일을 반환합니다")
def download_classified_excel(session_id: str):
    file_path = os.path.join(RESULT_DIR, f"{session_id}_classified.xlsx")
    if not os.path.exists(file_path):
        raise HTTPException(status_code=404, detail="엑셀 파일이 존재하지 않습니다.")

    return FileResponse(
        path=file_path,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=f"{session_id}_classified.xlsx"
    )
