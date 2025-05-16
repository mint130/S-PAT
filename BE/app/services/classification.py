import asyncio
from datetime import datetime
import io
import json
import logging
import os
import re
from typing import Any, Dict, List
from langchain_openai import OpenAIEmbeddings
import pandas as pd
from langchain_community.vectorstores import FAISS
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from langchain_core.prompts import ChatPromptTemplate
from langchain.output_parsers import PydanticOutputParser
from langchain_core.output_parsers import StrOutputParser
from langchain_core.runnables import RunnablePassthrough

from app.schemas.classification import ClassificationSchema, Patent
from app.schemas.message import Message, Progress

# 로그
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

embeddings_openai = OpenAIEmbeddings(model="text-embedding-3-large", dimensions=3072)

# 특허 분류 작업을 백그라운드에서 실행하는 함수
async def process_patent_classification(
    session_id: str,
    llm,
    retriever,
    df: pd.DataFrame,
    redis,
    progress_queue: asyncio.Queue
):
    logger.info(f"호출 완료")
    try:
        logger.info(f"시작")
        patents = []
        total_patents = len(df)
        
        # 로그 추가
        start_time = datetime.now()
        logger.info(f"[{session_id}] 분류 시작 시간: {start_time}")
        
        # 각 행에 대해 RAG 처리 및 분류 추가
        for index, row in df.iterrows():
            # 출원번호와 특허 정보 확인
            application_number = row.get('출원번호', f"KR10-XXXX-{index:07d}")  # 출원번호가 없을 경우 임의 생성
            title = row.get('특허명', row.get('발명의 명칭', ''))
            abstract = row.get('요약', '')
            
            # 특허 정보가 없으면 다음 행으로
            if not title and not abstract:
                continue
                
            # 특허 정보 구성
            patent_info = f"특허명: {title} 요약: {abstract}"
            
            # RAG를 통한 분류
            classifications = await classify_patent(llm, retriever, patent_info)
            
            # 원본 데이터프레임에 분류 결과 추가
            df.loc[index, '대분류코드'] = classifications["majorCode"]
            df.loc[index, '중분류코드'] = classifications["middleCode"]
            df.loc[index, '소분류코드'] = classifications["smallCode"]
            df.loc[index, '대분류명칭'] = classifications["majorTitle"]
            df.loc[index, '중분류명칭'] = classifications["middleTitle"]
            df.loc[index, '소분류명칭'] = classifications["smallTitle"]
            
            # 특허 정보 구성 (요청된 JSON 형식에 맞게)
            patent_data = Patent(
                applicationNumber=application_number,
                title=title,
                abstract=abstract,
                majorCode=classifications["majorCode"],
                middleCode=classifications["middleCode"],
                smallCode=classifications["smallCode"],
                majorTitle=classifications["majorTitle"],
                middleTitle=classifications["middleTitle"],
                smallTitle=classifications["smallTitle"]
            )
            
            # 1. 로그 출력
            logger.info(f"[{session_id}] 분류된 특허: {patent_data.model_dump()}")

            # 2. redis에 저장
            redis.rpush(session_id, patent_data.model_dump_json())

            # 3. 리스트 추가
            patents.append(patent_data)
            
            # 4. 진행률 업데이트
            progress_data = Progress(
                current=index + 1,
                total=total_patents,
                percentage=round(((index + 1) / total_patents) * 100, 2),
            )
            await progress_queue.put(progress_data.model_dump())  

            # time.sleep(0.5)
            await asyncio.sleep(0.5)
            
        # 세션 데이터 만료 시간 설정 (24시간)
        redis.expire(session_id, 86400)  # 86400초 = 24시간
        
        # 종료 시간 기록
        end_time = datetime.now()
        elapsed = end_time - start_time
        logger.info(f"[{session_id}] 분류 종료 시간: {end_time}")
        logger.info(f"[{session_id}] 분류 소요 시간: {elapsed} (총 {elapsed.total_seconds():.2f}초)")

        # 임시 파일명 생성
        filename = f"{session_id}_classified.xlsx"
        save_path = f"./classified_excels/{filename}"

        # 저장 디렉토리 확인
        os.makedirs(os.path.dirname(save_path), exist_ok=True)

        # 먼저 데이터프레임을 엑셀로 저장
        buffer = io.BytesIO()
        with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name="Sheet1")
        buffer.seek(0)

        # 헤더 노란색 처리
        wb = load_workbook(buffer)
        ws = wb.active
        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        for cell in ws[1]:
            cell.fill = yellow_fill

        # 저장
        wb.save(save_path)
        
        # 작업 완료 메시지
        completion_data = Message(
            status="completed",
            message="분류 작업이 완료되었습니다.",
        )

        await progress_queue.put(completion_data.model_dump())  
        
    except Exception as e:
        # 오류가 발생한 경우 메시지 전송
        
        error_message = Message(
            status="error",
            message=f"오류가 발생했습니다: {str(e)}"
        )
       
        await progress_queue.put(error_message.model_dump())

# 특허 정보를 바탕으로 분류하는 함수
async def classify_patent(llm, retriever, patent_info: str) -> Dict[str, str]:
    
    # 출력 파서 정의
    parser = PydanticOutputParser(pydantic_object=ClassificationSchema)
    
    # 프롬프트 템플릿 정의
    template = """
    다음은 특허 분류 데이터베이스에서 검색된 유사한 특허 정보입니다:
    
    {context}
    
    위 참고 정보를 바탕으로, 다음 특허 정보에 대한 대분류, 중분류, 소분류를 제공해주세요.
    대분류 코드와 명칭, 중분류 코드와 명칭, 소분류 코드와 명칭을 모두 포함해야 합니다.
    꼭 context안에 있는 기준으로 결과를 내세요.
    해당하는게 없으면 미분류로 결과 내세요.
    이유를 적지말고 응답 형식만 맞게 답 하세요.
    
    특허 정보: {query}
    
    너는 다음 JSON 형태로만 응답해야 해: {format_instructions}
    """


    prompt = ChatPromptTemplate.from_template(template)
    
        # 메타데이터를 포함한 형식으로 문서 포맷팅
    def format_docs(docs):
        formatted_docs = []
        for doc in docs:
            metadata = doc.metadata
            
            # 문서 레벨에 따라 다르게 포맷팅
            if metadata.get('level') == '중분류':
                formatted_doc = f"""
분류 정보:
- 코드: {metadata.get('code', '')}
- 명칭: {metadata.get('name', '')}
- 레벨: {metadata.get('level', '')}
- 상위 코드: {metadata.get('parent_code', '')}
- 상위 명칭: {metadata.get('parent_name', '')}
- 설명: {metadata.get('description', '')}

{doc.page_content}
"""
            elif metadata.get('level') == '소분류':
                formatted_doc = f"""
분류 정보:
- 코드: {metadata.get('code', '')}
- 명칭: {metadata.get('name', '')}
- 레벨: {metadata.get('level', '')}
- 상위 코드: {metadata.get('parent_code', '')}
- 상위 명칭: {metadata.get('parent_name', '')}
- 최상위 코드: {metadata.get('grand_parent_code', '')}
- 최상위 명칭: {metadata.get('grand_parent_name', '')}
- 설명: {metadata.get('description', '')}

{doc.page_content}
"""
            else:
                # 기타 레벨이나 메타데이터가 없는 경우
                formatted_doc = doc.page_content
                
            formatted_docs.append(formatted_doc)
            
        return "\n\n---\n\n".join(formatted_docs)
    
    rag_chain = (
        {
            "context": retriever | format_docs,
            "query": RunnablePassthrough(),
            "format_instructions": lambda _: parser.get_format_instructions()
        }
        | prompt
        | llm
        | StrOutputParser()
    )
    
    # 체인 실행
    result = await rag_chain.ainvoke(patent_info)
    
    # 결과 파싱
    try:
        cleaned = re.sub(r"```(?:json)?\s*([\s\S]+?)\s*```", r"\1", result.strip())
        parsed = json.loads(cleaned)

        return {
            "majorCode": parsed.get("majorCode", "미분류"),
            "majorTitle": parsed.get("majorTitle", "미분류"),
            "middleCode": parsed.get("middleCode", "미분류"),
            "middleTitle": parsed.get("middleTitle", "미분류"),
            "smallCode": parsed.get("smallCode", "미분류"),
            "smallTitle": parsed.get("smallTitle", "미분류"),
        }

    except Exception as e:
        print(f"분류 결과 파싱 중 오류: {str(e)}")
        # 파싱 실패 시 기본값 반환
        return {
            "majorCode": "미분류",
            "middleCode": "미분류",
            "smallCode": "미분류",
            "majorTitle": "미분류",
            "middleTitle": "미분류",
            "smallTitle": "미분류"
        }


def process_standards_for_vectordb(standards: List[Dict[Any, Any]]) -> List[Dict[Any, Any]]:
    """
    분류 체계 데이터를 벡터 DB에 저장하기 위한 형태로 가공합니다.
    """
    # 대분류, 중분류, 소분류 딕셔너리 생성
    high_categories = {}  # 대분류
    mid_categories = {}   # 중분류
    low_categories = {}   # 소분류
    
    # 먼저 각 분류 레벨별로 분류
    for standard in standards:
        level = standard['level']
        
        if level == '대분류':
            high_categories[standard['code']] = standard
        elif level == '중분류':
            mid_categories[standard['code']] = standard
        elif level == '소분류':
            low_categories[standard['code']] = standard
    
    # 처리된 문서를 저장할 리스트
    documents = []
    
    # 중분류 처리
    for code, mid_category in mid_categories.items():
        # 상위 대분류 찾기
        for high_code, high_category in high_categories.items():
            # 대분류 코드가 중분류 코드의 접두어인지 확인 (예: H01이 H01-01의 접두어)
            if code.startswith(high_code):
                # 임베딩 텍스트 생성 (대분류 설명 + 중분류 설명)
                embedding_text = f"{high_category['description']} {mid_category['description']}"
                
                # 벡터 DB용 문서 생성
                document = {
                    'code': code,
                    'name': mid_category['name'],
                    'level': mid_category['level'],
                    'parent_code': high_code,
                    'parent_name': high_category['name'],
                    'text': embedding_text,  # 임베딩에 사용될 텍스트
                    'description': mid_category['description']
                }
                
                documents.append(document)
                break
    
    # 소분류 처리
    for code, low_category in low_categories.items():
        # 상위 대분류와 중분류 찾기
        parent_mid = None
        parent_high = None
        
        # 먼저 상위 중분류 찾기
        for mid_code, mid_category in mid_categories.items():
            if code.startswith(mid_code):
                parent_mid = mid_category
                
                # 중분류의 상위 대분류 찾기
                for high_code, high_category in high_categories.items():
                    if mid_code.startswith(high_code):
                        parent_high = high_category
                        break
                
                break
        
        # 상위 분류를 모두 찾았는지 확인
        if parent_mid and parent_high:
            # 임베딩 텍스트 생성 (대분류 설명 + 중분류 설명 + 소분류 설명)
            embedding_text = f"{parent_high['description']} {parent_mid['description']} {low_category['description']}"
            
            # 벡터 DB용 문서 생성
            document = {
                'code': code,
                'name': low_category['name'],
                'level': low_category['level'],
                'parent_code': parent_mid['code'],
                'parent_name': parent_mid['name'],
                'grand_parent_code': parent_high['code'],
                'grand_parent_name': parent_high['name'],
                'text': embedding_text,  # 임베딩에 사용될 텍스트
                'description': low_category['description']
            }
            
            documents.append(document)
    
    return documents


def create_faiss_database(documents: List[Dict[Any, Any]], session_id: str, embedding_model_name: str = "text-embedding-ada-002"):
    """
    처리된 문서로부터 FAISS 벡터 데이터베이스를 생성하고 메모리에 저장합니다.
    """
    
    # FAISS에 사용할 텍스트와 메타데이터 준비
    texts = [doc['text'] for doc in documents]
    metadatas = documents  # 전체 문서를 메타데이터로 사용
    
    # FAISS 벡터 데이터베이스 생성
    vectordb = FAISS.from_texts(
        texts=texts,
        embedding=embeddings_openai,
        metadatas=metadatas
    )
    
    return vectordb
