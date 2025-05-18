from datetime import datetime
import io
import json
import logging
import os
import pickle
import re
import time
from typing import Any, Dict
from openai import RateLimitError as OpenAIRateLimitError
from anthropic import RateLimitError as ClaudeRateLimitError
from langchain_openai import OpenAIEmbeddings
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import pandas as pd
import requests
from app.core.celery import celery_app
from langchain_core.prompts import ChatPromptTemplate
from langchain.output_parsers import PydanticOutputParser
from langchain_core.output_parsers import StrOutputParser
from langchain_core.runnables import RunnablePassthrough
from langchain_community.vectorstores import FAISS
from app.core.redis import get_redis_client
from app.schemas.classification import ClassificationSchema, Patent
from app.core.llm import gpt, claude, gemini, grok
from app.schemas.message import Message, Progress
from google.api_core import exceptions

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# 유사도 평가 기준점 설정
THRESHOLD = 0.5  # 단일 임계값으로 변경

# LLM을 가져오는 함수
def get_llm_by_name(name: str):
    name = name.upper()
    if name == "GPT": 
        return gpt
    elif name == "CLAUDE": 
        return claude
    elif name == "GEMINI": 
        return gemini
    elif name == "GROK": 
        return grok
    else:
        raise ValueError(f"지원하지 않는 LLM 이름입니다: {name}")

# retriever을 가져오는 함수
def load_retriever_from_redis(session_id: str):
    redis = get_redis_client()
    path = redis.get(f"vectorstore:{session_id}:path")
    if path is None:
        raise Exception(f"{session_id}에 해당하는 벡터 저장소가 없습니다.")
    path = path.decode("utf-8") if isinstance(path, bytes) else path

    logger.info(f"FAISS index 경로: {path}")  # 경로 출력해서 확인

    embeddings = OpenAIEmbeddings(model="text-embedding-3-large", dimensions=3072)

    vector_store = FAISS.load_local(path, embeddings, allow_dangerous_deserialization=True)

    return vector_store.as_retriever(search_kwargs={"k": 3})


@celery_app.task(bind=True, retry_backoff=True, retry_backoff_max=10, retry_kwargs={'max_retries': 12})
def classify_patent(
    self,
    LLM: str, 
    session_id: str, 
    patent_info: str,
    application_number: str, 
    isAdmin: bool
 ) -> Dict[str, str]:
    redis = get_redis_client()
    logger.info("celery 시작")
    # retriever 가져옴
    retriever = load_retriever_from_redis(session_id)
    
    # llm 가져옴
    llm = get_llm_by_name(LLM)

    # 출력 파서 정의
    parser = PydanticOutputParser(pydantic_object=ClassificationSchema)
    
    # 프롬프트 템플릿 정의
    template = """
    다음은 특허 분류 데이터베이스에서 검색된 유사한 특허 정보입니다:
    
    {context}
    
    위 참고 정보를 바탕으로, 다음 특허 정보에 대한 대분류, 중분류, 소분류를 제공해주세요.
    대분류 코드와 명칭, 중분류 코드와 명칭, 소분류 코드와 명칭을 모두 포함해야 합니다.
    꼭 context안에 있는 기준으로 결과를 내세요.
    해당하는게 없으면 미분류로 결과 내세요.
    이유를 적지말고 응답 형식만 맞게 답 하세요.
    
    특허 정보: {query}
    
    너는 다음 JSON 형태로만 응답해야 해: {format_instructions}
    """
    prompt = ChatPromptTemplate.from_template(template)
    
    # 메타데이터를 포함한 형식으로 문서 포맷팅
    def format_docs(docs):
        formatted_docs = []
        for doc in docs:
            metadata = doc.metadata
            
            # 문서 레벨에 따라 다르게 포맷팅
            if metadata.get('level') == '중분류':
                formatted_doc = f"""
분류 정보:
- 코드: {metadata.get('code', '')}
- 명칭: {metadata.get('name', '')}
- 레벨: {metadata.get('level', '')}
- 상위 코드: {metadata.get('parent_code', '')}
- 상위 명칭: {metadata.get('parent_name', '')}
- 설명: {metadata.get('description', '')}

{doc.page_content}
"""
            elif metadata.get('level') == '소분류':
                formatted_doc = f"""
분류 정보:
- 코드: {metadata.get('code', '')}
- 명칭: {metadata.get('name', '')}
- 레벨: {metadata.get('level', '')}
- 상위 코드: {metadata.get('parent_code', '')}
- 상위 명칭: {metadata.get('parent_name', '')}
- 최상위 코드: {metadata.get('grand_parent_code', '')}
- 최상위 명칭: {metadata.get('grand_parent_name', '')}
- 설명: {metadata.get('description', '')}

{doc.page_content}
"""
            else:
                # 기타 레벨이나 메타데이터가 없는 경우
                formatted_doc = doc.page_content
                
            formatted_docs.append(formatted_doc)
            
        return "\n\n---\n\n".join(formatted_docs)
    
    rag_chain = (
        {
            "context": retriever | format_docs,
            "query": RunnablePassthrough(),
            "format_instructions": lambda _: parser.get_format_instructions()
        }
        | prompt
        | llm
        | StrOutputParser()
    )
    
    try:
        # 체인 실행
        logger.info(f"[{session_id}] rag_chain.invoke 시작")
        result = rag_chain.invoke(patent_info)
        logger.info(f"[{session_id}] rag_chain.invoke 완료")

        # 결과 파싱
        cleaned = re.sub(r"```(?:json)?\s*([\s\S]+?)\s*```", r"\1", result.strip())
        parsed = json.loads(cleaned)
        def replace_na(value):
            return value if value and value != "N/A" else "미분류"

        classifications = {
            "applicationNumber": application_number,  # 식별자 추가
            "majorCode": replace_na(parsed.get("majorCode")),
            "majorTitle": replace_na(parsed.get("majorTitle")),
            "middleCode": replace_na(parsed.get("middleCode")),
            "middleTitle": replace_na(parsed.get("middleTitle")),
            "smallCode": replace_na(parsed.get("smallCode")),
            "smallTitle": replace_na(parsed.get("smallTitle")),
        }
        
        # 진행률 업데이트
        if not isAdmin:
            progress_key = f"{session_id}:progress"
            current = redis.incr(f"{session_id}:progress_counter")  # +1 누적
            total = int(redis.get(f"{session_id}:total_count") or 1)
            percentage = int(current / total * 100)

            progress = Progress(
                current=current,
                total=total,
                percentage=percentage
            )

            redis.set(progress_key, progress.model_dump_json())
            redis.publish(progress_key, progress.model_dump_json())
            redis.expire(progress_key, 86400)

        return classifications
    
    except OpenAIRateLimitError as e:
        error_str = str(e)
        if 'rate_limit_exceeded' in error_str:
            # "Please try again in X.XXXs" 에서 시간 파싱
            logger.info(f"[{session_id}] OpenAI rate limit 발생")
            logger.info(f"[{session_id}] Retrying {self.request.retries}/12 due to OpenAI rate limit")
            raise self.retry(countdown= 5, max_retries = 12)
        else:
            raise

    except ClaudeRateLimitError as e:
        error_str = str(e)
        logger.info(f"[{session_id}] Claude rate limit 발생")
        raise self.retry(countdown= 5)
    
    except requests.exceptions.HTTPError as e:
        if e.response.status_code == 429:
            logger.info(f"[{session_id}] Grok rate limit 발생")
            # 헤더에서 재시도 시간 확인
            retry_after = int(e.response.headers.get('Retry-After', 5))
            
            # X-RateLimit 헤더 정보 로깅
            limit = e.response.headers.get('X-RateLimit-Limit-Tokens', 'unknown')
            remaining = e.response.headers.get('X-RateLimit-Remaining-Tokens', 'unknown')
            reset = e.response.headers.get('X-RateLimit-Reset-Tokens', 'unknown')
            
            logger.info(f"[{session_id}] Grok Rate Limit 정보 - 한도: {limit}, 남은 토큰: {remaining}, 초기화: {reset}초 후")
            
            # 명시적 대기 시간이 있으면 해당 시간만큼 대기 후 재시도
            if retry_after:
                logger.warning(f"[{session_id}] Grok 명시적 대기: {retry_after}초")
                time.sleep(retry_after)
            
            raise self.retry(countdown= retry_after)
        else:
            logger.error(f"[{session_id}] HTTP 요청 에러 발생: {e}")
            raise

    except exceptions.ResourceExhausted as e:
        # gemini limit error
        if e.status_code == 429:
            logger.info(f"[{session_id}] Gemini rate limit 발생")
            # 지수 백오프
            logger.warning(f"[{session_id}] 백오프")
            raise self.retry(countdown= 5)
        else:
            logger.error(f"[{session_id}] 다른 Gemini API 에러 발생: {e}")
            raise

    except Exception as e:
        logger.error(f"[{session_id}] 분류 결과 처리 중 오류 발생: {e}")
        return {
            "applicationNumber": application_number,
            "majorCode": "미분류",
            "middleCode": "미분류",
            "smallCode": "미분류",
            "majorTitle": "미분류",
            "middleTitle": "미분류",
            "smallTitle": "미분류"
        }

@celery_app.task
def classification_completion(results, session_id):
    try:
        redis = get_redis_client()
        time_key = f"{session_id}:time"
        # 종료 시간
        end_time = datetime.now()
        start_time = redis.get(time_key)
        start_time = float(start_time)
        start_time = datetime.fromtimestamp(start_time)
        
        # 시간 저장
        elapsed = end_time - start_time
        redis.set(time_key, elapsed.total_seconds())
        redis.expire(time_key, 86400)  # 86400초 = 24시간

        # df 불러오기
        with open(f"./temp_data/{session_id}.pkl", "rb") as f:
            df = pickle.load(f)

        for result in results:
            logger.info(f"분류 결과: {result}")
    
            application_number = result["applicationNumber"]  # 식별자 추출
        
            # 데이터프레임에서 해당 application_number를 가진 행 찾기
            matching_rows = df[df['출원번호'] == application_number]

            if not matching_rows.empty:
                index = matching_rows.index[0]
                row = df.loc[index]

                title = row.get('특허명', row.get('발명의 명칭', ''))
                abstract = row.get('요약', '')

                # 원본 데이터프레임에 분류 결과 추가
                df.loc[index, '대분류코드'] = result["majorCode"]
                df.loc[index, '중분류코드'] = result["middleCode"]
                df.loc[index, '소분류코드'] = result["smallCode"]
                df.loc[index, '대분류명칭'] = result["majorTitle"]
                df.loc[index, '중분류명칭'] = result["middleTitle"]
                df.loc[index, '소분류명칭'] = result["smallTitle"]

                # 특허 정보 구성
                patent_data = Patent(
                    applicationNumber=application_number,
                    title=title,
                    abstract=abstract,
                    majorCode=result["majorCode"],
                    middleCode=result["middleCode"],
                    smallCode=result["smallCode"],
                    majorTitle=result["majorTitle"],
                    middleTitle=result["middleTitle"],
                    smallTitle=result["smallTitle"]
                )
                redis.rpush(session_id, patent_data.model_dump_json())
                redis.expire(session_id, 86400)
            else:
                logger.warning(f"출원번호 {application_number}에 해당하는 행을 찾을 수 없음")

        # 임시 파일명 생성
        filename = f"{session_id}_classified.xlsx"
        save_path = f"./classified_excels/{filename}"

        # 저장 디렉토리 확인
        os.makedirs(os.path.dirname(save_path), exist_ok=True)

        # 먼저 데이터프레임을 엑셀로 저장
        buffer = io.BytesIO()
        with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name="Sheet1")
        buffer.seek(0)

        # 헤더 노란색 처리
        wb = load_workbook(buffer)
        ws = wb.active
        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        for cell in ws[1]:
            cell.fill = yellow_fill

        # 저장
        wb.save(save_path)

        # 알림
        message = Message(
            status="completed",
            message="분류 작업이 완료되었습니다."
        )
        redis.publish(f"{session_id}:progress", message.model_dump_json())
        redis.set(f"{session_id}:progress", message.model_dump_json())
        redis.expire(f"{session_id}:progress", 86400)

    except Exception as e:
        logger.error(f"[{session_id}] 분류 결과 처리 중 오류 발생: {e}")
        message = Message(
            status="error",
            message=f"분류 작업 중 오류가 발생했습니다: {str(e)}"
        )
        redis.publish(f"{session_id}:progress", message.model_dump_json())
        redis.set(f"{session_id}:progress", message.model_dump_json())
        return False
    
@celery_app.task
def evaluate_classification_by_vector(classification_result, LLM, session_id, patent_info, application_number)-> Dict[str, Any]:
    # 벡터 유사도 평가
    redis = get_redis_client()

    logger.info(f"벡터 유사도 평가 실행, 특허 분류 결과: {classification_result}")
    key = f"{session_id}:{LLM}"
    redis.hset(f"{key}:classifications", application_number, json.dumps(classification_result, ensure_ascii=False))
    
    # 1. 특허 정보와 가장 유사한 분류 체계 검색 (유사도 점수 포함)
    retriever = load_retriever_from_redis(session_id)
    vectorstore = retriever.vectorstore
    
    similar_docs_with_score = vectorstore.similarity_search_with_relevance_scores(patent_info, k=1)
    best_match, best_score = similar_docs_with_score[0]

    # numpy 타입을 Python 기본 타입으로 변환하고 0~1 사이로 정규화
    best_score = float(best_score)
    normalized_score = (best_score + 1) / 2  # -1~1 범위를 0~1 범위로 변환
    
    # 2. 평가 로직
    # LLM이 미분류로 분류한 경우
    is_unclassified = (
        classification_result["majorCode"] == "미분류" or
        classification_result["middleCode"] == "미분류" or
        classification_result["smallCode"] == "미분류"
    )
    
    # 평가 결과 결정
    if is_unclassified:
        # 미분류로 분류한 경우
        if normalized_score < THRESHOLD:
            # 유사도가 기준점보다 낮으면 미분류가 정확
            is_correct = True
            reason = f"유사도({normalized_score:.2f})가 기준점({THRESHOLD}) 미만이므로 미분류가 정확함"
        elif normalized_score == THRESHOLD:
            # 유사도가 기준점보다 낮으면 미분류가 정확
            is_correct = True
            reason = "분류 결과가 일치함"
        else:
            # 유사도가 기준점보다 높으면 미분류가 부정확
            is_correct = False
            reason = f"유사도({normalized_score:.2f})가 기준점({THRESHOLD}) 이상이므로 분류가 필요함"
    else:
        # 일반 분류의 경우
        if normalized_score < THRESHOLD:
            # 유사도가 기준점보다 낮으면 분류가 부정확
            is_correct = False
            reason = f"유사도({normalized_score:.2f})가 기준점({THRESHOLD}) 미만이므로 미분류여야 함"
        else:
            # 유사도가 기준점보다 높으면 분류 결과 비교
            is_correct = (
                classification_result["majorCode"] == best_match.metadata.get("grand_parent_code", "미분류") and
                classification_result["middleCode"] == best_match.metadata.get("parent_code", "미분류") and
                classification_result["smallCode"] == best_match.metadata.get("code", "미분류")
            )
            reason = "분류 결과가 일치함" if is_correct else "분류 결과가 일치하지 않음"
    
    # 3. 평가 결과 구성
    level = best_match.metadata.get("level", "")
    
    # 레벨에 따라 적절한 메타데이터 매핑
    if level == "대분류":
        major_code = best_match.metadata.get("code", "미분류")
        major_title = best_match.metadata.get("name", "미분류")
        middle_code = "미분류"
        middle_title = "미분류"
        small_code = "미분류"
        small_title = "미분류"
    elif level == "중분류":
        major_code = best_match.metadata.get("parent_code", "미분류")
        major_title = best_match.metadata.get("parent_name", "미분류")
        middle_code = best_match.metadata.get("code", "미분류")
        middle_title = best_match.metadata.get("name", "미분류")
        small_code = "미분류"
        small_title = "미분류"
    elif level == "소분류":
        major_code = best_match.metadata.get("grand_parent_code", "미분류")
        major_title = best_match.metadata.get("grand_parent_name", "미분류")
        middle_code = best_match.metadata.get("parent_code", "미분류")
        middle_title = best_match.metadata.get("parent_name", "미분류")
        small_code = best_match.metadata.get("code", "미분류")
        small_title = best_match.metadata.get("name", "미분류")
    else:
        major_code = "미분류"
        major_title = "미분류"
        middle_code = "미분류"
        middle_title = "미분류"
        small_code = "미분류"
        small_title = "미분류"

    return {
        "similarity_score": normalized_score,
        "best_match": {
            "majorCode": major_code,
            "majorTitle": major_title,
            "middleCode": middle_code,
            "middleTitle": middle_title,
            "smallCode": small_code,
            "smallTitle": small_title
        },
        "llm_classification": classification_result,
        "evaluation": {
            "is_correct": is_correct,
            "reason": reason
        }
    }

@celery_app.task(bind=True, retry_backoff=True, retry_backoff_max=10, retry_kwargs={'max_retries': 12})
def evaluate_classification_by_reasoning(
    self, 
    classification_result, 
    LLM, 
    session_id, 
    patent_info, 
    application_number):
    # LLM 평가
    redis = get_redis_client()

    logger.info(f"LLM 평가 실행, 특허 분류 결과: {classification_result}")
    key = f"{session_id}:{LLM}"
    redis.hset(f"{key}:classifications", application_number, json.dumps(classification_result, ensure_ascii=False))
  
    # 유사도 기반 분류 체계 검색
    retriever = load_retriever_from_redis(session_id)
    similar_docs = retriever.invoke(patent_info)
    # 유사도 기반 분류 체계 포맷팅
    similar_classifications = []
    for doc in similar_docs:
        metadata = doc.metadata
        classification = f"""
분류 체계:
- 대분류: {metadata.get('grand_parent_code', '')} ({metadata.get('grand_parent_name', '')})
- 중분류: {metadata.get('parent_code', '')} ({metadata.get('parent_name', '')})
- 소분류: {metadata.get('code', '')} ({metadata.get('name', '')})
- 설명: {metadata.get('description', '')}
"""
        similar_classifications.append(classification)
    
    similar_classifications_text = "\n---\n".join(similar_classifications)

    # 프롬프트 템플릿 정의
    template = """
    당신은 특허 분류 전문가입니다. 다음 특허의 분류 결과를 평가해주세요.
    특허 정보는 대상 특허 데이터의 특허명과 요약을 포함합니다.
    현재 분류 결과는 저장된 분류 체계를 기반으로 분류한 결과입니다.
    유사도 기반 추천 분류 체계는 분류 체계 중 특허 정보와 가장 유사한 3개의 분류입니다.
    당신의 목표는 이것들을 기준으로 이 특허 분류가 정확한지 평가하는 것입니다.
    평가 요구사항과 주의사항에 맞게 잘 평가해 주세요.
    
    [특허 정보]
    {patent_info}

    [현재 분류 결과]
    대분류: {major_code} ({major_title})
    중분류: {middle_code} ({middle_title})
    소분류: {small_code} ({small_title})

    [유사도 기반 추천 분류 체계]
    {similar_classifications}

    [평가 요구사항]
    1. 분석: 유사도 기반 추천 분류 체계와 현재 분류 결과를 비교하여 평가해주세요.
    2. 점수: 0.0부터 1.0 사이의 점수로 평가해주세요.
       - 0.0: 완전히 부적절한 분류
       - 0.5: 부분적으로 적절한 분류
       - 1.0: 완벽하게 적절한 분류
    3. 이유: 분석을 1줄 요약하여 작성해주세요.

    [주의사항]
    1. 반드시 분류 체계를 기반으로 평가해야 합니다.
    2. 분류 체계에 없는 분류는 '미분류'로 간주합니다.
    - 미분류의 예시: 특허 정보가 제시된 분류 체계만으로 분류할 수 없다고 판단될 경우, 미분류로 판단합니다.
    3. 평가 요구 사항의 포맷 [1. 분석 2. 점수 3. 이유]를 반드시 지킵니다.
    """

    prompt = ChatPromptTemplate.from_template(template)
    
    # 프롬프트에 값 채우기
    filled_prompt = prompt.format(
        patent_info=patent_info,
        major_code=classification_result["majorCode"],
        major_title=classification_result["majorTitle"],
        middle_code=classification_result["middleCode"],
        middle_title=classification_result["middleTitle"],
        small_code=classification_result["smallCode"],
        small_title=classification_result["smallTitle"],
        similar_classifications=similar_classifications_text
    )
    try:
        response = claude.invoke(filled_prompt)
        
        # 응답에서 점수 추출 (0.0~1.0 사이의 숫자)
        try:
            score_match = re.search(r'0\.\d+|1\.0', response.content)
            if score_match:
                score = float(score_match.group())
            else:
                score = 0.0
                logger.warning("점수를 찾을 수 없어 기본값 0.0을 사용합니다.")
        except Exception as e:
            logger.error(f"점수를 추출 중 오류 발생: {str(e)}")
            score = 0.0

        # 응답에서 평가 이유 추출
        reason_match = re.search(r"이유.*?[:：\s]\s*(.+?)(?:\n##|$)", response.content, re.DOTALL)
        reason = reason_match.group(1).strip() if reason_match else ""

        return {
            "score": score,
            "reason": reason
        } 
    
    except ClaudeRateLimitError as e:
        logger.info(f"[{session_id}] Claude rate limit 발생")
        raise self.retry(countdown= 5)

@celery_app.task
def collect_evaluation_results(evaluation_results, LLM, session_id, application_number):
    # 진행도 계산, 두 결과 합침
    
    logger.info(f"진행도 계산, 두결과 합침, 평가 결과: {evaluation_results}")

    redis = get_redis_client()
    vector_result = evaluation_results[0]
    reasoning_result = evaluation_results[1]

    key = f"{session_id}:{LLM}"
    
    # 평가 점수와 이유 저장
    redis.hset(f"{key}:reason", application_number, json.dumps(reasoning_result, ensure_ascii=False))
  
    # 진행도 계산
    progress_key = f"{key}:progress"
    current = redis.incr(f"{key}:progress_counter")  # +1 누적
    total = int(redis.get(f"{key}:total_count") or 1)
    percentage = int(current / total * 100)

    progress = Progress(
        current=current,
        total=total,
        percentage=percentage
    )

    redis.set(progress_key, progress.model_dump_json())
    redis.publish(progress_key, progress.model_dump_json())
    redis.expire(progress_key, 86400)
    
    return {
        "vector_based": vector_result,
        "reasoning": reasoning_result
    }

def calculate_vector_based_score(patent_evaluations):
    total_patents = len(patent_evaluations)
    
    # 벡터 기반 평가 점수 계산
    vector_correct_classifications = sum(
        1 for eval in patent_evaluations 
        if eval["vector_based"]["evaluation"]["is_correct"] == True
    )
    vector_accuracy = (vector_correct_classifications / total_patents) * 100
    
    # Reasoning LLM 평가 점수 계산 (평균 점수)
    reasoning_scores = [
        eval["reasoning"]["score"] 
        for eval in patent_evaluations
    ]
    reasoning_average_score = sum(reasoning_scores) / total_patents * 100
    
    # 간소화된 평가 점수만 반환
    return {
        "total_patents": total_patents,
        "vector_accuracy": vector_accuracy,
        "reasoning_score": reasoning_average_score
    }

@celery_app.task
def evaluation_completion(results, LLM, session_id):
    # 모든 처리 끝남
    logger.info(f"모든 처리 끝남 {results}")
    try:
        redis = get_redis_client()
        key = f"{session_id}:{LLM}"

        time_key = f"{key}:time"
        
        # 종료 시간
        end_time = datetime.now()
        start_time = redis.get(time_key)
        start_time = float(start_time)
        start_time = datetime.fromtimestamp(start_time)
        
        # 시간 저장
        elapsed = end_time - start_time
        redis.set(time_key, elapsed.total_seconds())
        redis.expire(time_key, 86400)  # 86400초 = 24시간

        # df 불러오기
        with open(f"./temp_data/{session_id}.pkl", "rb") as f:
            df = pickle.load(f)

        # df에 맞게 patents에 특허 저장, reasoning 다시
        for index, row in df.iterrows():
            # 출원번호와 특허 정보 확인
            application_number = str(row.get('출원번호', f"KR10-XXXX-{index:07d}"))
            title = str(row.get('특허명', row.get('발명의 명칭', '')))
            abstract = str(row.get('요약', ''))
            classifications_json = redis.hget(f"{key}:classifications", application_number)
            if classifications_json:
                classifications = json.loads(classifications_json)
            
            # 특허 정보 구성
            patent_data = Patent(
                applicationNumber=application_number,
                title=title,
                abstract=abstract,
                majorCode=classifications["majorCode"],
                middleCode=classifications["middleCode"],
                smallCode=classifications["smallCode"],
                majorTitle=classifications["majorTitle"],
                middleTitle=classifications["middleTitle"],
                smallTitle=classifications["smallTitle"]
            )

            # 평가 결과 리스트 형식으로 저장
            logger.info(f"{application_number} 평가 결과 저장")
            redis.rpush(f"{key}:patents", patent_data.model_dump_json())
            
            # 평가 이유 리스트 형식으로 저장
            logger.info(f"{application_number} 평가 이유 저장")
            reason_json = redis.hget(f"{key}:reason", application_number)
            redis.rpush(f"{key}:reasoning", reason_json)

        # 임시로 저장한 데이터 삭제
        redis.delete(f"{key}:reason")
        redis.delete(f"{key}:classifications")

        # 최종 평가 점수 계산 (간소화된 버전)
        evaluation_score = calculate_vector_based_score(results)

        # Redis에 간소화된 평가 점수만 저장
        simplified_evaluation = {
            "vector_accuracy": evaluation_score["vector_accuracy"],
            "reasoning_score": evaluation_score["reasoning_score"]
        }
        redis.set(f"{key}:evaluation", json.dumps(simplified_evaluation, ensure_ascii=False))

        # 평과 결과 만료 시간 설정
        redis.expire(f"{key}:evaluation", 86400)
        redis.expire(f"{key}:patents", 86400)
        redis.expire(f"{key}:reasoning", 86400)
        
        # 알림
        message = Message(
            status="completed",
            message="분류 및 평가 작업이 완료되었습니다."
        )

    except Exception as e:
        logger.error(f"[{session_id}:{LLM}] 분류 결과 처리 중 오류 발생: {e}")
        message = Message(
            status="error",
            message=f"분류 작업 중 오류가 발생했습니다: {str(e)}"
        )
        redis.publish(f"{key}:progress", message.model_dump_json())
        redis.set(f"{key}:progress", message.model_dump_json())
        return False